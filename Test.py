import pandas as pd
'''
df = pd.DataFrame([ [-10, -9, 8], 
[6, 2, -4], 
[-8, 5, 1]], 
columns=['a', 'b', 'c'])
'''
dict = {'Column1':['q','w','e','r','t','y'],
        'Column2':['a','s','d','f','g','h'],
        'Column3':[1,2,3,4,5,6]}
#print(dict)
df = pd.DataFrame(dict)
print('')
#print (df, '\n')
#print('df.shape \n',  df.shape[0], '------',  df.shape[1], '\n', df.shape)
#print('df.info() \n', type(df.info()))

df_test = df.copy()
df_test['Column4'] = 0
#print(df_test)
df_test.loc[2,'Column3'] = 100
#df_test[df_test.loc[:,'Column3'] == 4] = 100   # Заменить даннные во всех столбцах в той строке, где в столбце4 стоит 4, на 100
#df_test[df_test.Column3 == 4] = 100     # Аналог предыдущей строки: # Заменить даннные во всех столбцах в той строке, где в столбце4 стоит 4, на 100
df_test.loc[(df_test.Column3 == 4), 'Column3']  = 100
print(df_test)

df_test.to_excel('test.xlsx', index=False)

import openpyxl

'''from openpyxl import Workbook
wb = Workbook()
ws = wb.active
print('Workbook()', type(wb))
print('wb.active', type(ws))
ws['A1'] = 5
print(ws['A1'].value) '''

from openpyxl import load_workbook
wb = load_workbook('test.xlsx')  # открыть существующую книгу Excel
ws = wb.active
a = ws['B3'].value
for temp in ws:
    print(temp[2], 'равно', temp[2].value)
    if(temp[2].value == 100): temp[3].value = "Yes"
    print(temp[3], 'равно', temp[3].value)
wb.save('test.xlsx')

#df_test.sort_values(by = ["Column3"], ascending=True, inplace=True, na_position='last', ignore_index=True)

#xl = pd.ExcelFile('C:\Users\NewSokrat\Project_NalogKKM\test.xlsx')

#print(df_test.describe())  #информация о числовых столбцах (int, float)
#print(df_test.describe(include=['0']))   #информация о текстовых столбцах (object) - НЕ РАБОТАЕТ!!!
#print(df.describe(include = 'all')) # статистика включает не только числовые столбцы, но и строки